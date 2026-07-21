#!/usr/bin/env python3
"""
==============================================================================
 pipeline_metricas.py  ·  Easier-Notes SQA
==============================================================================
 Capa de PROCESAMIENTO del ecosistema de aseguramiento.

 Flujo:
   Herramientas (xUnit, Newman, JMeter)
        -> datos crudos (.trx, cobertura.xml, .jtl, reportes)
        -> [ Python + Pandas ]  limpia, estructura, calcula métricas
        -> Excel de auditoría (.xlsx)   <- capa de validación humana
        -> datos.json                   <- alimenta el dashboard

 Uso:
   python pipeline_metricas.py                # usa datos de ejemplo si no hay crudos
   python pipeline_metricas.py --dir ./resultados

 Produce:
   - Auditoria_Metricas_EasierNotes.xlsx
   - datos.json  (compatible con el dashboard)
==============================================================================
"""
import argparse
import glob
import json
import os
import sys
import xml.etree.ElementTree as ET
from datetime import datetime

import pandas as pd


# ==========================================================================
# 1) LECTURA DE DATOS CRUDOS  (con Pandas y parseo XML)
#    Cada función devuelve un DataFrame o un dict. Si no encuentra el
#    archivo, devuelve None para poder caer a los datos de ejemplo.
# ==========================================================================

def leer_trx(carpeta):
    """Lee los .trx de `dotnet test` y devuelve un DataFrame por proyecto."""
    filas = []
    ns = {"t": "http://microsoft.com/schemas/VisualStudio/TeamTest/2010"}
    for trx in glob.glob(os.path.join(carpeta, "**", "*.trx"), recursive=True):
        try:
            root = ET.parse(trx).getroot()
            c = root.find(".//t:ResultSummary/t:Counters", ns)
            if c is not None:
                filas.append({
                    "archivo": os.path.basename(trx),
                    "total": int(c.get("total", 0)),
                    "passed": int(c.get("passed", 0)),
                    "failed": int(c.get("failed", 0)),
                })
        except Exception as e:
            print(f"  (aviso) no pude leer {trx}: {e}")
    return pd.DataFrame(filas) if filas else None


def leer_cobertura(carpeta):
    """Lee cobertura.cobertura.xml (coverlet) -> (linea%, rama%)."""
    for cov in glob.glob(os.path.join(carpeta, "**", "*cobertura*.xml"), recursive=True):
        try:
            root = ET.parse(cov).getroot()
            return (round(float(root.get("line-rate", 0)) * 100, 1),
                    round(float(root.get("branch-rate", 0)) * 100, 1))
        except Exception as e:
            print(f"  (aviso) no pude leer {cov}: {e}")
    return None


def leer_jtl(ruta):
    """Lee un .jtl de JMeter con Pandas y calcula las métricas de desempeño."""
    if not os.path.exists(ruta):
        return None
    try:
        df = pd.read_csv(ruta)
        # JMeter usa 'elapsed' (ms) y 'success' (true/false)
        elapsed = df["elapsed"].astype(int)
        n = len(elapsed)
        errores = (df["success"].astype(str).str.lower() == "false").sum()
        return {
            "muestras": int(n),
            "p95": int(elapsed.quantile(0.95)),
            "promedio": int(elapsed.mean()),
            "mediana": int(elapsed.median()),
            "max": int(elapsed.max()),
            "min": int(elapsed.min()),
            "throughput": round(n / (elapsed.sum() / 1000), 2) if elapsed.sum() else 0,
            "errorPct": round(errores / n * 100, 2) if n else 0,
        }
    except Exception as e:
        print(f"  (aviso) no pude leer {ruta}: {e}")
        return None


# ==========================================================================
# 2) DATOS DE EJEMPLO  (los reales medidos en el proyecto)
#    Se usan cuando no hay archivos crudos, para que el pipeline
#    siempre produzca una salida demostrable.
# ==========================================================================

def datos_ejemplo():
    return {
        "pruebas": pd.DataFrame([
            {"archivo": "EasierNotes.Tests.trx",            "total": 65, "passed": 65, "failed": 0},
            {"archivo": "EasierNotes.IntegrationTests.trx", "total": 32, "passed": 32, "failed": 0},
        ]),
        "cobertura": (98.8, 92.1),
        "jm01": {"muestras": 400, "p95": 14, "promedio": 19, "mediana": 9,
                 "max": 1222, "min": 6, "throughput": 3.0, "errorPct": 1.0},
        "asercionesApi": 44,
    }


# ==========================================================================
# 3) CONSTRUCCIÓN DE LOS DATASETS ESTRUCTURADOS (Pandas)
#    Aquí Pandas "limpia y estructura de forma reproducible".
# ==========================================================================

def construir_datasets(pruebas_df, cobertura, jm01, aserciones):
    total = int(pruebas_df["total"].sum())
    passed = int(pruebas_df["passed"].sum())
    failed = int(pruebas_df["failed"].sum())
    linea, rama = cobertura

    # --- Dataset de métricas de PROCESO ---
    proceso = pd.DataFrame([
        {"Métrica": "Total de pruebas de código", "Valor": total,           "Unidad": "pruebas",  "Umbral": "-",   "Estado": "OK"},
        {"Métrica": "Pruebas superadas",          "Valor": passed,          "Unidad": "pruebas",  "Umbral": total, "Estado": "OK" if failed == 0 else "REVISAR"},
        {"Métrica": "Tasa de éxito",              "Valor": round(passed/total*100, 1) if total else 0, "Unidad": "%", "Umbral": 100, "Estado": "OK" if failed == 0 else "REVISAR"},
        {"Métrica": "Cobertura de línea",         "Valor": linea,           "Unidad": "%",        "Umbral": 80,    "Estado": "OK" if linea >= 80 else "BAJO"},
        {"Métrica": "Cobertura de rama",          "Valor": rama,            "Unidad": "%",        "Umbral": 80,    "Estado": "OK" if rama >= 80 else "BAJO"},
        {"Métrica": "Aserciones de API",          "Valor": aserciones,      "Unidad": "aserciones","Umbral": "-",  "Estado": "OK"},
    ])

    # --- Dataset de métricas de PRODUCTO ---
    producto = pd.DataFrame([
        {"Métrica": "Latencia p95 (creación)", "Valor": jm01["p95"],       "Unidad": "ms",  "Umbral": 500, "Estado": "OK" if jm01["p95"] <= 500 else "EXCEDE"},
        {"Métrica": "Latencia promedio",       "Valor": jm01["promedio"],  "Unidad": "ms",  "Umbral": "-", "Estado": "info"},
        {"Métrica": "Latencia mediana",        "Valor": jm01["mediana"],   "Unidad": "ms",  "Umbral": "-", "Estado": "info"},
        {"Métrica": "Latencia máxima",         "Valor": jm01["max"],       "Unidad": "ms",  "Umbral": "-", "Estado": "frío"},
        {"Métrica": "Throughput",              "Valor": jm01["throughput"],"Unidad": "req/s","Umbral": "-","Estado": "info"},
        {"Métrica": "Porcentaje de error",     "Valor": jm01["errorPct"],  "Unidad": "%",   "Umbral": 0,   "Estado": "OK" if jm01["errorPct"] <= 1 else "REVISAR"},
    ])

    # --- Dataset por nivel de prueba ---
    niveles = pd.DataFrame([
        {"Nivel": "Unitarias",   "Cantidad": 65, "Enfoque": "Caja blanca", "Estado": "verde"},
        {"Nivel": "Integración", "Cantidad": 32, "Enfoque": "Caja gris",   "Estado": "verde"},
        {"Nivel": "Caja negra",  "Cantidad": 18, "Enfoque": "Caja negra",  "Estado": "verde"},
        {"Nivel": "Sistema",     "Cantidad": 17, "Enfoque": "Caja negra",  "Estado": "verde"},
        {"Nivel": "Aceptación",  "Cantidad": 9,  "Enfoque": "Caja negra",  "Estado": "parcial"},
    ])

    return proceso, producto, niveles


def construir_desempeno(jm01, jm02, jm03):
    """Estructura las métricas de los 3 planes JMeter en un DataFrame.
    Cada plan puede ser None (aún no ejecutado); se marca como 'pendiente'."""
    planes = [
        ("JM-01", "Comportamiento Temporal", jm01, 500),
        ("JM-02", "Capacidad",               jm02, None),
        ("JM-03", "Utilización de Recursos", jm03, None),
    ]
    filas = []
    for plan, sub, d, umbral in planes:
        if d is None:
            filas.append({"Plan": plan, "Subcaracterística": sub, "Muestras": "-",
                          "p95 (ms)": "-", "Promedio (ms)": "-", "Máx (ms)": "-",
                          "Throughput (req/s)": "-", "Error (%)": "-", "Estado": "pendiente"})
        else:
            if umbral is not None:
                estado = "OK" if d["p95"] <= umbral else "EXCEDE"
            else:
                estado = "medido"
            filas.append({"Plan": plan, "Subcaracterística": sub, "Muestras": d["muestras"],
                          "p95 (ms)": d["p95"], "Promedio (ms)": d["promedio"], "Máx (ms)": d["max"],
                          "Throughput (req/s)": d["throughput"], "Error (%)": d["errorPct"], "Estado": estado})
    return pd.DataFrame(filas)


# ==========================================================================
# 4) EXPORTAR EXCEL DE AUDITORÍA  (capa de validación humana)
# ==========================================================================

def exportar_excel(proceso, producto, niveles, desempeno, meta, ruta_xlsx):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    AZUL = "1D3F72"; ORO = "E6C56B"; VERDE = "2E7D32"; ROJO = "C0392B"; GRIS = "F1EFE8"
    thin = Side(style="thin", color="BBBBBB")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()

    def escribir_hoja(ws, df, titulo):
        ws.sheet_view.showGridLines = False
        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        c = ws.cell(row=1, column=1, value=titulo)
        c.font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 26
        # Encabezados
        for j, col in enumerate(df.columns, start=1):
            hc = ws.cell(row=3, column=j, value=col)
            hc.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            hc.fill = PatternFill("solid", fgColor=AZUL)
            hc.alignment = Alignment(horizontal="center", vertical="center")
            hc.border = borde
        # Datos
        for i, (_, row) in enumerate(df.iterrows(), start=4):
            for j, val in enumerate(row, start=1):
                dc = ws.cell(row=i, column=j, value=val)
                dc.font = Font(name="Arial", size=10)
                dc.border = borde
                dc.alignment = Alignment(horizontal="center" if j > 1 else "left", vertical="center")
                if i % 2 == 0:
                    dc.fill = PatternFill("solid", fgColor=GRIS)
                # Colorear la columna Estado
                if df.columns[j-1] == "Estado":
                    if str(val) in ("OK",):
                        dc.font = Font(name="Arial", size=10, bold=True, color=VERDE)
                    elif str(val) in ("REVISAR", "EXCEDE", "BAJO"):
                        dc.font = Font(name="Arial", size=10, bold=True, color=ROJO)
        # Anchos
        for j, col in enumerate(df.columns, start=1):
            ancho = max(14, min(40, int(df[col].astype(str).str.len().max()) + 6, len(col) + 8))
            ws.column_dimensions[ws.cell(row=3, column=j).column_letter].width = ancho

    # Hoja 1: Resumen / portada de auditoría
    ws0 = wb.active
    ws0.title = "Resumen"
    ws0.sheet_view.showGridLines = False
    ws0.merge_cells("A1:D1")
    t = ws0.cell(row=1, column=1, value="Auditoría de Métricas · Easier-Notes")
    t.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=AZUL)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws0.row_dimensions[1].height = 30
    info = [
        ("Proyecto", meta["proyecto"]),
        ("Característica", meta["caracteristica"]),
        ("Generado por", "Pipeline Python + Pandas"),
        ("Fecha de proceso", meta["actualizado"]),
        ("Propósito", "Validación y auditoría de métricas antes de la visualización"),
    ]
    for i, (k, v) in enumerate(info, start=3):
        kc = ws0.cell(row=i, column=1, value=k)
        kc.font = Font(name="Arial", size=10, bold=True, color=AZUL)
        vc = ws0.cell(row=i, column=2, value=v)
        vc.font = Font(name="Arial", size=10)
    ws0.column_dimensions["A"].width = 20
    ws0.column_dimensions["B"].width = 55
    # Nota de auditoría
    ws0.merge_cells(start_row=9, start_column=1, end_row=9, end_column=4)
    nota = ws0.cell(row=9, column=1, value="Revise la columna 'Estado': OK (verde) indica conformidad; "
                                           "REVISAR/EXCEDE/BAJO (rojo) requieren atención antes de publicar al dashboard.")
    nota.font = Font(name="Arial", size=9, italic=True, color="666666")
    nota.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.row_dimensions[9].height = 30

    # Hojas de métricas
    escribir_hoja(wb.create_sheet("Métricas de proceso"), proceso, "Métricas de PROCESO — salud del aseguramiento")
    escribir_hoja(wb.create_sheet("Métricas de producto"), producto, "Métricas de PRODUCTO — Eficiencia de Desempeño")
    escribir_hoja(wb.create_sheet("Niveles de prueba"), niveles, "Distribución por nivel de prueba")
    escribir_hoja(wb.create_sheet("Desempeño JMeter"), desempeno, "Desempeño — 3 planes JMeter (subcaracterísticas)")

    wb.save(ruta_xlsx)
    print(f"  [OK] Excel de auditoría -> {ruta_xlsx}")


# ==========================================================================
# 5) GENERAR datos.json  (alimenta el dashboard)
# ==========================================================================

def generar_json(proceso, producto, niveles, desempeno, jm01, jm02, jm03, meta, ruta_json):
    def val(df, metrica):
        r = df[df["Métrica"] == metrica]
        return r.iloc[0]["Valor"] if len(r) else None

    def bloque_jm(d, nombre):
        if d is None:
            return {"nombre": nombre, "estado": "pendiente"}
        return {"nombre": nombre, "muestras": d["muestras"], "p95": d["p95"],
                "promedio": d["promedio"], "mediana": d["mediana"], "max": d["max"],
                "min": d["min"], "throughput": d["throughput"], "errorPct": d["errorPct"]}

    data = {
        "meta": meta,
        "kpis": {
            "totalPruebas": int(val(proceso, "Total de pruebas de código")),
            "pruebasVerdes": int(val(proceso, "Pruebas superadas")),
            "coberturaLinea": float(val(proceso, "Cobertura de línea")),
            "coberturaRama": float(val(proceso, "Cobertura de rama")),
            "asercionesApi": int(val(proceso, "Aserciones de API")),
            "hallazgos": 4,
            "p95": int(val(producto, "Latencia p95 (creación)")),
            "umbralP95": 500,
            "throughput": float(val(producto, "Throughput")),
            "errorPct": float(val(producto, "Porcentaje de error")),
        },
        "niveles": [
            {"nombre": r["Nivel"], "cantidad": int(r["Cantidad"]),
             "tipo": r["Enfoque"], "estado": r["Estado"], "enfoque": r["Enfoque"]}
            for _, r in niveles.iterrows()
        ],
        "cobertura": {"linea": float(val(proceso, "Cobertura de línea")),
                      "rama": float(val(proceso, "Cobertura de rama")), "umbral": 80},
        "desempeno": {
            "jm01": bloque_jm(jm01, "JM-01 · Comportamiento Temporal"),
            "jm02": bloque_jm(jm02, "JM-02 · Capacidad"),
            "jm03": bloque_jm(jm03, "JM-03 · Utilización de Recursos"),
        },
        "proceso_tabla": proceso.to_dict(orient="records"),
        "producto_tabla": producto.to_dict(orient="records"),
        "desempeno_tabla": desempeno.to_dict(orient="records"),
    }
    with open(ruta_json, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  [OK] datos.json -> {ruta_json}")


# ==========================================================================
# MAIN
# ==========================================================================

def main():
    ap = argparse.ArgumentParser(description="Pipeline de métricas SQA (Python+Pandas -> Excel -> JSON)")
    ap.add_argument("--dir", default=".", help="Carpeta con los resultados crudos (.trx, .jtl, cobertura)")
    ap.add_argument("--out", default=".", help="Carpeta de salida")
    args = ap.parse_args()

    print("=" * 60)
    print(" Pipeline de métricas SQA · Easier-Notes")
    print("=" * 60)

    # --- Intentar leer datos reales; si no hay, usar ejemplo ---
    print("\n[1/4] Leyendo datos crudos con Pandas...")
    pruebas_df = leer_trx(args.dir)
    cobertura = leer_cobertura(args.dir)
    jm01 = leer_jtl(os.path.join(args.dir, "jm01.jtl"))
    jm02 = leer_jtl(os.path.join(args.dir, "jm02.jtl"))
    jm03 = leer_jtl(os.path.join(args.dir, "jm03.jtl"))
    aserciones = 44  # de los reportes de Newman

    ej = datos_ejemplo()
    if pruebas_df is None:
        print("      (sin .trx) usando datos de ejemplo para pruebas")
        pruebas_df = ej["pruebas"]
    if cobertura is None:
        print("      (sin cobertura.xml) usando datos de ejemplo")
        cobertura = ej["cobertura"]
    if jm01 is None:
        print("      (sin jm01.jtl) usando datos de ejemplo para JM-01")
        jm01 = ej["jm01"]
    if jm02 is None:
        print("      (sin jm02.jtl) JM-02 queda como pendiente")
    if jm03 is None:
        print("      (sin jm03.jtl) JM-03 queda como pendiente")

    meta = {
        "proyecto": "Easier-Notes",
        "caracteristica": "Eficiencia de Desempeño",
        "subcaracteristicas": ["Comportamiento Temporal", "Utilización de Recursos", "Capacidad"],
        "norma": "ISO/IEC/IEEE 29119 · ISO/IEC 25010 · IEEE 1061",
        "actualizado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "version": "1.0",
    }

    print("\n[2/4] Estructurando datasets con Pandas...")
    proceso, producto, niveles = construir_datasets(pruebas_df, cobertura, jm01, aserciones)
    desempeno = construir_desempeno(jm01, jm02, jm03)
    print(proceso.to_string(index=False))
    print("\n  Desempeño (3 planes JMeter):")
    print(desempeno.to_string(index=False))

    print("\n[3/4] Exportando Excel de auditoría...")
    exportar_excel(proceso, producto, niveles, desempeno, meta, os.path.join(args.out, "Auditoria_Metricas_EasierNotes.xlsx"))

    print("\n[4/4] Generando datos.json para el dashboard...")
    generar_json(proceso, producto, niveles, desempeno, jm01, jm02, jm03, meta, os.path.join(args.out, "datos.json"))

    print("\n" + "=" * 60)
    print(" Pipeline completado. Flujo: crudos -> Pandas -> Excel -> JSON -> Dashboard")
    print("=" * 60)


if __name__ == "__main__":
    main()
